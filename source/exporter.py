"""Exports a report."""

from datetime import date
import os
import time
import tkinter as tk
from tkinter import ttk
from tkinter.messagebox import showinfo, showerror
import shutil
import openpyxl
import PIL

from iconer import set_window_icon


class ReportExporter(tk.Toplevel):
    """Docstring"""
    def __init__(self, parent, project, results_queue=None):
        """Instantiate the core."""
        tk.Toplevel.__init__(self, parent)
        self.results_queue = results_queue
        if len(self.results_queue) == 0:
            self.withdraw()
            showinfo(parent=self, message="You must evaulate a set of data before exporting a report.")
            self.destroy()
            return
        self.attributes('-topmost', 'true')
        self.title('Project details')
        set_window_icon(self)
        self.parent = parent
        self.results_queue = results_queue

        self.header_details = None
        self.project = project
        self.build(project)

    def build(self, project):
        """Build all the widgets."""
        vcmd = (self.register(self.is_numeric))

        container = tk.Frame(self)
        details_container = tk.LabelFrame(container, text="Report header")
        anal_lbl = tk.Label(details_container, text="Analysis number:", anchor='w')
        comp_lbl = tk.Label(details_container, text="Company:", anchor='w')
        sample_lbl = tk.Label(details_container, text="Sample point:", anchor='w')
        cust_lbl = tk.Label(details_container, text="Customer:", anchor='w')
        client_lbl = tk.Label(details_container, text="Submitted by:", anchor='w')
        date_lbl = tk.Label(details_container, text="Date submitted:", anchor='w')
        temp_lbl = tk.Label(details_container, text="Test temperature (°F):", anchor='w')
        cl_lbl = tk.Label(details_container, text="Chlorides (mg/L):", anchor='w')
        bicarb_lbl = tk.Label(details_container, text="Bicarbonates (mg/L):", anchor='w')
        bicarb_adj_lbl = tk.Label(details_container, text="Bicarbonate adjustment (mg/L):", anchor='w')

        anal_ent = ttk.Entry(details_container, width=25)
        comp_ent = ttk.Entry(details_container, width=25)
        sample_ent = ttk.Entry(details_container, width=25)
        cust_ent = ttk.Entry(details_container, width=25)
        client_ent = ttk.Entry(details_container, width=25)
        date_ent = ttk.Entry(details_container, width=25)
        temp_ent = ttk.Spinbox(details_container, width=23, from_=150, to=220, validate='all', validatecommand=(vcmd, '%P'))
        cl_ent = ttk.Spinbox(details_container, width=23, from_=0, to=999999, validate='all', validatecommand=(vcmd, '%P'))
        bicarb_ent = ttk.Spinbox(details_container, width=23, from_=0, to=999999, validate='all', validatecommand=(vcmd, '%P'))
        bicarb_adj_ent = ttk.Spinbox(details_container, width=23, from_=0, to=999999, validate='all', validatecommand=(vcmd, '%P'))

        header_widgets = (anal_ent, comp_ent, sample_ent, cust_ent, client_ent, date_ent, temp_ent, cl_ent, bicarb_ent, bicarb_adj_ent)

        anal_lbl.grid(row=0, column=0, sticky='e', pady=1, padx=2)
        comp_lbl.grid(row=1, column=0, sticky='e', pady=1, padx=2)
        sample_lbl.grid(row=2, column=0, sticky='e', pady=1, padx=2)
        cust_lbl.grid(row=3, column=0, sticky='e', pady=1, padx=2)
        client_lbl.grid(row=4, column=0, sticky='e', pady=1, padx=2)
        date_lbl.grid(row=5, column=0, sticky='e', pady=1, padx=2)
        temp_lbl.grid(row=6, column=0, sticky='e', pady=1, padx=2)
        cl_lbl.grid(row=7, column=0, sticky='e', pady=1, padx=2)
        bicarb_lbl.grid(row=8, column=0, sticky='e', pady=1, padx=2)
        bicarb_adj_lbl.grid(row=9, column=0, sticky='e', pady=1, padx=2)

        anal_ent.grid(row=0, column=1, sticky='e', pady=1, padx=2)
        comp_ent.grid(row=1, column=1, sticky='e', pady=1, padx=2)
        sample_ent.grid(row=2, column=1, sticky='e', pady=1, padx=2)
        cust_ent.grid(row=3, column=1, sticky='e', pady=1, padx=2)
        client_ent.grid(row=4, column=1, sticky='e', pady=1, padx=2)
        date_ent.grid(row=5, column=1, sticky='e', pady=1, padx=2)
        temp_ent.grid(row=6, column=1, sticky='e', pady=1, padx=2)
        cl_ent.grid(row=7, column=1, sticky='e', pady=1, padx=2)
        bicarb_ent.grid(row=8, column=1, sticky='e', pady=1, padx=2)
        bicarb_adj_ent.grid(row=9, column=1, sticky='e', pady=1, padx=2)

        details_container.pack(pady=2, padx=2, fill='both')

        water_quals = tk.LabelFrame(container, text="Water quality")
        water_qual_ents = []
        clarities = ["Clear", "Slightly Hazy", "Hazy", "Other"]
        for i, trial in enumerate(self.results_queue['result_titles']):
            tk.Label(water_quals, text=trial).grid(row=i, column=0, sticky='e')
            ent = ttk.Combobox(water_quals, values=clarities, justify='center')
            ent.grid(row=i, column=1)
            ent.set(clarities[0])
            water_qual_ents.append(ent)
        water_quals.pack(pady=2, padx=2)

        submit_btn = ttk.Button(container, text="Confirm", command=lambda: self.get_details(header_widgets, water_qual_ents), width=25)

        submit_btn.pack(pady=2, padx=2)
        container.pack()

        # insert defaults
        anal_ent.insert(0, "#-#")
        temp_ent.insert(0, 200)
        self.project = project.split('\\')
        if len(self.project) >= 3:
            comp_ent.insert(0, self.project[-1].split('-')[0].strip())
            sample_ent.insert(0, self.project[-1].split('-')[1].strip())
            cust_ent.insert(0, self.project[-2].strip())

    def get_details(self, widgets, water_qual_ents):
        """Docstring."""
        details = [widget.get() for widget in widgets]
        details = [value.strip() for value in details]
        trial_clarities = [widget.get() for widget in water_qual_ents]
        trial_clarities = [value.strip() for value in trial_clarities]
        self.header_details = details
        self.results_queue['trial_clarities'] = trial_clarities
        self.export_report()

    def export_report(self):
        """Export the reporter's results_queue to an .xlsx file."""
        template_path = self.parent.parser.get('report settings', 'template path')
        if not os.path.isfile(template_path):
            showerror(
                parent=self,
                message="No valid template file found.\nYou can set the template path in Settings > Report Settings."
            )
            return
        print("Preparing export")
        start = time.time()
        analysis_no = self.header_details[0]
        company = self.header_details[1]
        sample = self.header_details[2]
        customer = self.header_details[3]
        client = self.header_details[4]
        sub_date = self.header_details[5]

        def ret_num(str) -> int:
            str = str.replace(",", "")
            if str == "":
                str = 0
            return round(float(str))

        temp = ret_num(self.header_details[6])
        cl = ret_num(self.header_details[7])
        bicarbs = ret_num(self.header_details[8])
        bicarb_adj = ret_num(self.header_details[9])

        file = f"{analysis_no.replace(' ', '')} {self.project[-1]} CaCO3 Scale Block Analysis.xlsx"
        report_path = os.path.join(self.parent.mainwin.project, file)

        print(f"Copying report template to\n{report_path}")
        shutil.copyfile(template_path, report_path)

        print(f"Populating file\n{report_path}")
        workbook = openpyxl.load_workbook(report_path)
        ws = workbook.active

        img_filename = f"{self.project[-1]}.png"
        img_path = os.path.join(self.parent.mainwin.project, img_filename)
        print("Making temp resized plot image")
        try:
            img = PIL.Image.open(img_path)
        except FileNotFoundError:
            print("Couldn't find the plot image file, aborting export")
            return
        img = img.resize((667, 257))
        img_path = img_filename[:-4]
        img_path += "- temp.png"
        img.save(img_path)
        img = openpyxl.drawing.image.Image(img_path)
        img.anchor = 'A28'
        ws._images[1] = img

        blank_times = self.results_queue['blank_times']
        result_titles = self.results_queue['result_titles']
        result_values = self.results_queue['result_values']
        durations = self.results_queue['durations']
        baseline = self.results_queue['baseline']
        ylim = self.results_queue['ylim']
        max_psis = self.results_queue['max_psis']
        trial_clarities = self.results_queue['trial_clarities']

        brine_comp = f"Synthetic Field Brine, Chlorides = {cl:,} mg/L"
        if bicarb_adj != 0:
            brine_comp += f" (Bicarbs adjusted to {bicarbs:,} mg/L)"
        ws['D12'] = brine_comp

        ws['D10'] = f"{temp} °F"


        ws['C4'] = customer
        ws['C5'] = client
        ws['C6'] = company
        ws['C7'] = sample
        today = date.today().strftime("%B %d, %Y")
        ws['I4'] = analysis_no
        ws['I6'] = sub_date
        ws['I7'] = f"{today}"
        ws['D11'] = f"{baseline} psi"
        ws['G16'] = round(ylim)

        print(f"customer: {customer}")
        print(f"company: {company}")
        print(f"well / sample point: {sample}")

        blank_time_cells = [f"E{i}" for i in range(16, 18)]
        chem_name_cells = [f"A{i}" for i in range(19, 27)]
        chem_conc_cells = [f"D{i}" for i in range(19, 27)]
        duration_cells = [f"E{i}" for i in range(19, 27)]
        max_psi_cells = [f"G{i}" for i in range(19, 27)]
        protection_cells = [f"H{i}" for i in range(19, 27)]
        clarity_cells = [f"J{i}" for i in range(19, 27)]

        chem_names = [" ".join(title.split(' ')[:-2]) for title in result_titles]
        chem_concs = [" ".join(title.split(' ')[-2:-1]) for title in result_titles]

        for (cell, blank_time) in zip(blank_time_cells, blank_times):
            ws[cell] = round(blank_time / 60, 2)
        for (cell, name) in zip(chem_name_cells, chem_names):
            ws[cell] = f"{name}"
        for (cell, conc) in zip(chem_conc_cells, chem_concs):
            ws[cell] = round(float(conc), 1)
        for (cell, duration) in zip(duration_cells, durations):
            ws[cell] = round(float(duration), 2)
        for (cell, psi) in zip(max_psi_cells, max_psis):
            ws[cell] = round(psi)
        for (cell, score) in zip(protection_cells, result_values):
            score = float(score[:-1])
            if score >= 100:
                score = 100
            ws[cell] = score / 100  # the template has conditional % formatting
        for(cell, clarity) in zip(clarity_cells, trial_clarities):
            ws[cell] = clarity

        rows_with_data = [16, 17, *range(19, 27)]  # where the data is
        hide_rows = []  # rows we want to hide
        resize_rows = []  # rows we want to resize

        for i in rows_with_data:
            if ws[f'A{i}'].value is None:  # if the cell is empty
                hide_rows.append(i)  # add it to the list of rows to hide
            else:
                resize_rows.append(i)  # add it to the list of rows to reszie

        row_height = 200 / len(resize_rows)  # we have ~200px to work with total
        if row_height >= 30:  # we don't want any rows bigger than this
            row_height = 30
        for row in resize_rows:  # this does the resizing
            ws.row_dimensions[row].height = row_height
        for row in hide_rows:  # this hides the empty rows
            ws.row_dimensions[row].hidden = True

        print(f"Saving report to\n{report_path}")
        workbook.save(filename=report_path)
        print("Removing temp files")
        os.remove(img_path)
        print(f"Finished export in {round(time.time() - start, 2)} s")
        showinfo(
            parent=self,
            message=f"Report exported to\n{report_path}"
        )
        self.destroy()

    def is_numeric(self, val):
        """Validate that user input is numeric."""
        parts = val.split(',')
        results = []
        for chars in parts:
            results.append(chars == "" or str.isdigit(chars))
        return all(results)
