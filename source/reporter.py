"""Evaluates data and makes a plot."""

from datetime import date
import os  # handling file paths
import pickle  # storing plotter settings
import tkinter as tk  # GUI
from tkinter import ttk, filedialog, font  # type: ignore
from tkinter.messagebox import showinfo, showwarning, showerror
import shutil
import time
import PIL

import matplotlib as mpl
import matplotlib.pyplot as plt  # plotting the data
from matplotlib.ticker import MultipleLocator
import openpyxl
from pandas import Series, DataFrame, read_csv  # reading the data

import settings
from iconer import set_window_icon
from seriesentry import SeriesEntry
from evaluator import evaluate


class Reporter(tk.Toplevel):
    """Class for collecting data and exporting results from evaluator."""

    def __init__(self, parent, *args, **kwargs):
        """Init the reporter."""
        tk.Toplevel.__init__(self, parent, *args, **kwargs)
        set_window_icon(self)
        self.core = parent
        self.mainwin = self.core.mainwin
        self.parser = self.core.parser
        self.loc = tk.StringVar()
        self.resizable(0, 0)
        self.build()
        # if there's a project file in the dir go ahead and open it
        project = self.mainwin.project.split('\\')
        short_proj = project[-1]
        pickle_project_file = f"{short_proj}.pct"
        file = os.path.join(self.mainwin.project, pickle_project_file)
        if os.path.isfile(file):
            self.unpickle_plot(file)

    def build(self) -> None:
        """Make the widgets."""
        def_font = font.nametofont("TkDefaultFont")
        bold_font = font.Font(font=def_font)
        bold_font.config(weight='bold')
        self.winfo_toplevel().title("Report Generator")
        vcmd = (self.register(self.is_numeric))

        self.pltbar = tk.Menu(self)
        self.pltbar.add_command(label="Save project", command=lambda: self.pickle_plot())
        self.pltbar.add_command(label="Load project", command=lambda: self.unpickle_plot())
        self.pltbar.add_command(label='Export report', command=lambda: self.export_report())
        self.pltbar.add_command(label='Help', command=lambda: self.show_help())

        self.winfo_toplevel().configure(menu=self.pltbar)

        # a LabelFrame to hold the SeriesEntry widgets
        self.ent_frm = tk.LabelFrame(
            master=self,
            # NOTE: this is a dirty way of avoiding using many labels
            text=(
                "File path:                                 Series title:                                          Data to evaluate:"
            ),
            font=bold_font
        )
        default_pump = self.parser.get('test settings', 'default pump')
        for _ in range(10):
            SeriesEntry(self.ent_frm, default_pump).grid(padx=2)
        self.ent_frm.grid(row=0, padx=3, pady=2)

        # to hold the settings entries
        self.set_frm = tk.LabelFrame(master=self, text="Plot parameters", font=bold_font)
        # settings labels
        time_lbl = tk.Label(master=self.set_frm, text="Time limit (min):")
        fail_lbl = tk.Label(master=self.set_frm, text="Max pressure (psi):")
        base_lbl = tk.Label(master=self.set_frm, text="Baseline pressure (psi):")
        # grid the labels
        time_lbl.grid(row=0, column=0, sticky='w', padx=5)
        fail_lbl.grid(row=0, column=1, sticky='w', padx=5)
        base_lbl.grid(row=0, column=2, sticky='w', padx=5)
        # settings entries
        self.time_limit = ttk.Spinbox(
            self.set_frm,
            width=14,
            from_=0, to=1440,
            justify='center',
            validate='all', validatecommand=(vcmd, '%P')
        )
        self.fail_psi = ttk.Spinbox(
            self.set_frm,
            width=14,
            from_=0, to=5000,
            justify='center',
            validate='all', validatecommand=(vcmd, '%P')
        )
        self.baseline = ttk.Spinbox(
            self.set_frm,
            width=14,
            from_=0, to=5000,
            justify='center',
            validate='all', validatecommand=(vcmd, '%P')
        )
        # grid settings entries
        self.time_limit.grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.fail_psi.grid(row=1, column=1, sticky='w', padx=5, pady=2)
        self.baseline.grid(row=1, column=2, sticky='w', padx=5, pady=2)

        # insert default plot parameters
        self.time_limit.insert(0, self.parser.getint('test settings', 'time limit minutes'))
        self.fail_psi.insert(0, self.parser.getint('test settings', 'fail psi'))
        self.baseline.insert(0, self.parser.getint('test settings', 'default baseline'))

        # make a button to plot, then grid it
        self.pltbtn = ttk.Button(
            master=self.set_frm,
            text="Evaluate",
            width=30,
            command=lambda: self.make_plot(**self.prep_plot())
        )
        self.pltbtn.grid(row=2, columnspan=4, pady=2)
        # grid the settings frame
        self.set_frm.grid(row=1, pady=2)

    def prep_plot(self) -> dict:
        """Return a dict of lists.

        paths - str paths of original datafiles
        titles - str tuple of original series titles
        plotpumps - str tuple which pump's pressure to evaluate
        plot_params - a tuple of (str, int, int, (floats,))
                    last arg optional
        """
        print("Preparing plot from Reporter fields")
        entry_widgets = self.ent_frm.winfo_children()
        paths = [child.path.get() for child in entry_widgets]
        titles = [child.title.get() for child in entry_widgets]
        plotpumps = [child.plotpump.get() for child in entry_widgets]
        # look for empty entries in the form
        if self.time_limit.get() != '':
            xlim = int(self.time_limit.get())
        else:
            xlim = self.parser.getint('test settings', 'time limit minutes')
        if self.fail_psi.get() != '':
            ylim = int(self.fail_psi.get())
        else:
            ylim = self.parser.getint('test settings', 'fail psi')
        if self.baseline.get() != '':
            baseline = int(self.baseline.get())
        else:
            baseline = self.parser.getint('test settings', 'default baseline')

        plot_params = ('bmh', xlim, ylim, baseline)
        plot_data = {
            'paths': paths,
            'titles': titles,
            'plotpumps': plotpumps,
            'plot_params': plot_params
        }
        return plot_data

    def make_plot(self, paths, titles, plotpumps, plot_params) -> None:
        """Make a new plot of the data, then evaluate it."""
        print("Spawning a new plot \n")
        start = time.time()

        # give names to plot parameters
        style = plot_params[0]
        if plot_params[1] != '':
            xlim = plot_params[1]
        else:
            xlim = int(self.mainwin.timelim.get())

        if plot_params[2] != '':
            ylim = plot_params[2]
        else:
            ylim = int(self.mainwin.failpsi.get())

        with plt.style.context(style):
            colors = self.parser.get('report settings', 'color cycle').split(',')
            colors = [i.strip() for i in colors]
            try:
                mpl.rcParams['axes.prop_cycle'] = mpl.cycler(color=colors)
            except ValueError:
                showwarning(
                    parent=self,
                    title="Invalid Color Cycle",
                    message="Tried to use invalid colors, reverting to defaults"
                )
                colors = settings.DEFAULT_DICT['report settings']['color cycle']  # type: ignore
                mpl.rcParams['axes.prop_cycle'] = mpl.cycler(color=colors)
            fig, ax = plt.subplots(figsize=(12.5, 5), dpi=100)
            ax.set_xlabel("Time (min)")
            ax.set_xlim(left=0, right=xlim)
            ax.set_ylabel("Pressure (psi)")
            ax.set_ylim(top=ylim)
            ax.yaxis.set_major_locator(MultipleLocator(100))
            ax.grid(color='darkgrey', alpha=0.65, linestyle='-')
            ax.set_facecolor('w')
            fig.canvas.set_window_title(self.mainwin.winfo_toplevel().title())
            plt.tight_layout()

            blanks = []
            trials = []
            for path, title, plotpump in zip(paths, titles, plotpumps):
                if os.path.isfile(path):
                    df = DataFrame(read_csv(path))
                else:
                    df = DataFrame(
                        data={'Minutes': [0], 'Pump 1': [0], 'Pump 2': [0]})
                try:
                    df[plotpump]  # make sure the column exists
                except KeyError:  # backwards compat w/ old data file headers
                    plotpump = plotpump.replace("Pump", "PSI")
                # if the trial is a blank run change the line style
                if title == "":
                    pass
                elif "blank" in str(title).lower():
                    ax.plot(df['Minutes'], df[plotpump], label=title, linestyle=('-.'))
                    blanks.append(Series(df[plotpump], name=title))
                else:  # plot using default line style
                    ax.plot(df['Minutes'], df[plotpump], label=title)
                    trials.append(Series(df[plotpump], name=title))

            ax.legend(loc='best')
            baseline = plot_params[3]

            # some tests to validate user input
            if len(blanks) == 0 and len(trials) == 0:
                showwarning(
                    parent=self,
                    title="No data selected",
                    message="Click a 'File path:' entry to select a data file"
                )
            elif len(blanks) == 0:
                showwarning(
                    parent=self,
                    title="No series designated as blank",
                    message="At least one series title must contain 'blank'"
                )
            elif len(trials) == 0:
                showwarning(
                    parent=self,
                    title="No trial data selected",
                    message="Must select least one trial not titled 'blank'")
            else:
                self.get_results(blanks, trials, baseline, xlim, ylim)
                fig.show()
                project = self.mainwin.project.split('\\')
                short_proj = project[-1]
                image = f"{short_proj}.png"
                image_path = os.path.join(self.mainwin.project, image)
                fig.savefig(image_path)
                print(f"Saved plot image to\n{image_path}")
                print(f"Finished plotting in {round(time.time() - start, 2)} s")

    def pickle_plot(self) -> None:
        """Pickle a list to a file in the project directory."""
        project = self.mainwin.project.split('\\')
        short_proj = project[-1]
        _pickle = f"{short_proj}.pct"
        _path = os.path.join(self.mainwin.project, _pickle)
        with open(_path, 'wb') as file:
            print(f"Pickling project to\n{_path}")
            pickle.dump(self.prep_plot(), file=file)
        showinfo(
            parent=self,
            title="Saved successfully",
            message=f"Project data saved to\n{_path}"
        )

    def unpickle_plot(self, pickle_file = None) -> None:
        """Unpickle a list of string 3-tuples into SeriesEntry widgets."""
        if not pickle_file:  # if one isn't passed in
            file = filedialog.askopenfilename(
                initialdir="C:\"",
                title="Select data to plot:",
                filetypes=[("ScaleWiz project files", "*.pct")]
            )
        else:
            file = pickle_file
        # this puts data paths into their original entries
        if file != '':
            print(f"Unpickling project file\n{file}")
            with open(file, 'rb') as file:
                plot = pickle.load(file)
            into_widgets = zip(
                plot['paths'],
                plot['titles'],
                plot['plotpumps'],
                self.ent_frm.winfo_children()
            )

            print("Populating plot parameters")
            # plot_params are ('bmh', xlim, ylim, baseline)
            plot_params = (plot['plot_params'][1:])
            param_widgets = (self.time_limit, self.fail_psi, self.baseline)
            for widget, parameter in zip(param_widgets, plot_params):
                widget.delete(0, 'end')
                widget.insert(0, parameter)

            print("Populating series entry fields")
            for path, title, plotpump, widget in into_widgets:
                widget.path.delete(0, 'end')
                widget.path.insert(0, path)
                self.after(150, widget.path.xview_moveto, 1)
                widget.title.delete(0, 'end')
                widget.title.insert(0, title)
                self.after(150, widget.title.xview_moveto, 1)
                widget.plotpump.set(plotpump)
                #  NOTE: after for race condition
                #  https://stackoverflow.com/questions/29334544/

        # raise the settings window
        self.lift()

    def get_results(self, blanks, trials, baseline, xlim, ylim):
        """Evaluate the data."""
        print("Getting results from evaluator")
        interval = self.parser.getint('test settings', 'interval seconds')
        proj = self.mainwin.project
        self.results_queue, self.log = evaluate(
            proj, blanks, trials, baseline, xlim, ylim, interval
        )
        # results_queue is (blank_times, result_titles, result_values, durations, baseline, ylim, max_psis)

        project = self.mainwin.project.split('\\')[-1].strip()
        log_file = os.path.join(self.mainwin.project, f"{project} log.txt")
        with open(log_file, 'w') as file:
            file.write('\n'.join(self.log))
        print(f"Wrote calculations log to \n{log_file}\n")

        result_window = tk.Toplevel(self)
        result_window.attributes('-topmost', 'true')
        result_window.title("Results")
        set_window_icon(result_window)

        def_bg = result_window.cget('bg')
        tk.Label(
            master=result_window,
            text="Trial",
            anchor='w'
        ).grid(row=0, column=0, sticky='w', padx=35, pady=3)
        tk.Label(
            master=result_window,
            text="% protection",
            anchor='w'
        ).grid(row=0, column=1, sticky='w', padx=35, pady=3)

        for i, title in enumerate(self.results_queue[1]):
            entry = tk.Entry(result_window, bg=def_bg, width=len(title))
            entry.insert(0, title)
            entry.configure(state='readonly', relief='flat')
            entry.grid(row=i + 1, column=0, sticky='W', padx=35, pady=3)

        for i, value in enumerate(self.results_queue[2]):
            entry = tk.Entry(result_window, bg=def_bg, width=10)
            entry.insert(0, value)
            entry.configure(state='readonly', relief='flat')
            entry.grid(row=i + 1, column=1, sticky='W', padx=35, pady=3)

    def export_report(self):
        """Export the reporter's results_queue to an .xlsx file."""
        if not hasattr(self, 'results_queue'):
            showinfo(
                parent=self,
                message="You must evaulate a set of data before exporting a report."
            )
            return
        template_path = self.parser.get('report settings', 'template path')
        if not os.path.isfile(template_path):
            showerror(
                parent=self,
                message="No valid template file found.\nYou can set the template path in Settings > Report Settings."
            )
            return

        print("Preparing export")
        start = time.time()
        project = self.mainwin.project.split('\\')
        try:
            company = project[-1].split('-')[0].strip()
            sample_point = project[-1].split('-')[1].strip()
            customer = project[-2].strip()
            short_proj = project[-1].strip()
        except IndexError:
            company = project[-1]
            sample_point = company
            customer = company
            short_proj = company

        file = f"#-# {short_proj} Calcium Carbonate Scale Block Analysis.xlsx"
        report_path = os.path.join(self.mainwin.project, file)

        print(f"Copying report template to\n{report_path}")
        shutil.copyfile(template_path, report_path)

        print(f"Populating file\n{report_path}")
        workbook = openpyxl.load_workbook(report_path)
        ws = workbook.active

        img_filename = f"{short_proj}.png"
        img_path = os.path.join(self.mainwin.project, img_filename)
        print("Making temp resized plot image")
        try:
            img = PIL.Image.open(img_path)
            # img = img.resize((700, 265))
        except FileNotFoundError:
            print("Couldn't find the plot image file, aborting export")
            return
        img = img.resize((667, 257))
        img_path = img_filename[:-4]
        img_path += "- temp.png"
        img.save(img_path)
        img = openpyxl.drawing.image.Image(img_path)
        img.anchor = 'A28'
        ws._images[1] = img

        blank_times = self.results_queue[0]
        result_titles = self.results_queue[1]
        result_values = self.results_queue[2]
        durations = self.results_queue[3]
        baseline = self.results_queue[4]
        ylim = self.results_queue[5]
        max_psis = self.results_queue[6]

        # TODO:  move this mapping into the config somehow
        ws['C4'] = customer
        ws['C6'] = company
        ws['C7'] = sample_point
        today = date.today().strftime("%B %d, %Y")
        ws['I7'] = f"{today}"
        ws['D11'] = f"{baseline} psi"
        ws['G16'] = round(ylim)

        print(f"customer: {customer}")
        print(f"company: {company}")
        print(f"well / sample point: {sample_point}")

        blank_time_cells = [f"E{i}" for i in range(16, 18)]
        chem_name_cells = [f"A{i}" for i in range(19, 27)]
        chem_conc_cells = [f"D{i}" for i in range(19, 27)]
        duration_cells = [f"E{i}" for i in range(19, 27)]
        max_psi_cells = [f"G{i}" for i in range(19, 27)]
        protection_cells = [f"H{i}" for i in range(19, 27)]

        chem_names = [" ".join(title.split(' ')[:-2]) for title in result_titles]
        chem_concs = [" ".join(title.split(' ')[-2:-1]) for title in result_titles]

        for (cell, blank_time) in zip(blank_time_cells, blank_times):
            ws[cell] = round(blank_time / 60, 2)
        for (cell, name) in zip(chem_name_cells, chem_names):
            ws[cell] = f"{name}"
        for (cell, conc) in zip(chem_conc_cells, chem_concs):
            ws[cell] = round(float(conc), 1)
        for (cell, duration) in zip(duration_cells, durations):
            ws[cell] = round(float(duration), 2)
        for (cell, psi) in zip(max_psi_cells, max_psis):
            ws[cell] = round(psi)
        for (cell, score) in zip(protection_cells, result_values):
            score = float(score[:-1])
            if score >= 100:
                score = 100
            ws[cell] = score / 100  # the template has conditional % formatting

        rows_with_data = [16, 17, *range(19, 27)]  # where the data is
        hide_rows = []  # rows we want to hide
        resize_rows = []  # rows we want to resize

        for i in rows_with_data:
            if ws[f'A{i}'].value is None:  # if the cell is empty
                hide_rows.append(i)  # add it to the list of rows to hide
            else:
                resize_rows.append(i)  # add it to the list of rows to reszie

        row_height = 200 / len(resize_rows)  # we have ~200px to work with total
        if row_height >= 30:  # we don't want any rows bigger than this
            row_height = 30
        for row in resize_rows:  # this does the resizing
            ws.row_dimensions[row].height = row_height
        for row in hide_rows:  # this hides the empty rows
            ws.row_dimensions[row].hidden = True

        print(f"Saving report to\n{report_path}")
        workbook.save(filename=report_path)
        print("Removing temp files")
        os.remove(img_path)
        print(f"Finished export in {round(time.time() - start, 2)} s")
        showinfo(
            parent=self,
            message=f"Report exported to\n{report_path}"
        )

    def is_numeric(self, P):
        """Validate that user input is numeric."""
        if str.isdigit(P) or P == "":
            return True
        else:
            return False

    def show_help(self):
        """Show a help dialog to the user."""
        showinfo(
            parent=self,
            title="Help: Using the Report Generator",
            message="""
Clicking on a 'File path' field will prompt you to select the .csv file of the data you want to plot. You may change the default 'Series title' to appear on the plot.

Including the word 'blank' in a series title will designate it as such for the purposes of calculations, and will set the data to plot as a dashed line. You must select at least one blank and one non-blank set of data to evaluate results.

Save project: Saves the current Report Generator form to a .pct file

Load project: Repopulates the Report Generator form from a .pct file

Export report: Makes a copy of the file at 'Template Path' (see Report Settings) and populates it with the results of the evaluation.
"""
        )
